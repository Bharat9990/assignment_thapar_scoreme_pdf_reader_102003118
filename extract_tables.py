import fitz  # PyMuPDF for PDF processing
from openpyxl import Workbook
from openpyxl.utils.exceptions import IllegalCharacterError
import string
import sys

def extract_tables_from_pdf(pdf_file):
    doc = fitz.open(pdf_file)
    tables = []
    
    for page_num in range(len(doc)):
        page = doc.load_page(page_num)
        table_blocks = page.get_text("dict")["blocks"]
        
        for b in table_blocks:
            if b["type"] == 0 and "lines" in b:  # block containing lines (a table)
                table = []
                for line in b["lines"]:
                    row = []
                    for span in line["spans"]:
                        cleaned_text = sanitize_text(span["text"])
                        row.append(cleaned_text)
                    table.append(row)
                tables.append(table)
    
    doc.close()
    return tables

def sanitize_text(text):
    # Remove characters that cannot be used in Excel sheets
    valid_chars = string.printable
    cleaned_text = ''.join(c for c in text if c in valid_chars)
    return cleaned_text

def save_tables_to_excel(tables, excel_file):
    wb = Workbook()
    for idx, table in enumerate(tables, start=1):
        ws = wb.create_sheet(f"Sheet{idx}")
        for row_idx, row in enumerate(table, start=1):
            for col_idx, cell in enumerate(row, start=1):
                try:
                    ws.cell(row=row_idx, column=col_idx, value=cell)
                except IllegalCharacterError:
                    cleaned_cell = sanitize_text(cell)
                    ws.cell(row=row_idx, column=col_idx, value=cleaned_cell)
    
    wb.save(excel_file)


def main():
    if len(sys.argv) < 3:
        print("Usage: python script.py <pdf_file_path> <excel_file_path>")
        return
    
    pdf_file = sys.argv[1]
    excel_file = sys.argv[2]
    
    tables = extract_tables_from_pdf(pdf_file)
    save_tables_to_excel(tables, excel_file)
    
    print(f"Tables extracted from {pdf_file} and saved to {excel_file}")

if __name__ == "__main__":
    main()
